# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, jsonify, send_file
import os
import io
import re
import ssl
import json
import base64
import struct
import urllib.request
from urllib.parse import urlencode
from database import get_db, init_db, import_data
from config import BAIDU_API_KEY, BAIDU_SECRET_KEY, BAIDU_TOKEN_URL, BAIDU_OCR_URL

app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/search', methods=['POST'])
def search():
    data = request.get_json()
    valve_model = data.get('valve_model', '').strip()
    customer_name = data.get('customer_name', '').strip()

    if not valve_model and not customer_name:
        return jsonify({'success': False, 'message': '请输入阀体型号或客户名称'})

    conn = get_db()
    c = conn.cursor()

    query = """
        SELECT DISTINCT ps.id, ps.spec_number, ps.product_model, ps.description
        FROM packaging_specs ps
        JOIN spec_entries se ON ps.id = se.spec_id
        WHERE 1=1
    """
    params = []

    if valve_model:
        query += " AND se.valve_model LIKE ?"
        params.append(f"%{valve_model}%")

    if customer_name:
        query += " AND se.customer_name LIKE ?"
        params.append(f"%{customer_name}%")

    if valve_model and customer_name:
        query += """
            AND ps.id IN (
                SELECT spec_id FROM spec_entries
                WHERE valve_model LIKE ? AND customer_name LIKE ?
            )
        """
        params.extend([f"%{valve_model}%", f"%{customer_name}%"])

    c.execute(query, params)
    specs = c.fetchall()

    results = []
    for spec in specs:
        c.execute("""
            SELECT seq_number, customer_name, valve_model, special_note, remark
            FROM spec_entries WHERE spec_id = ?
        """, (spec['id'],))
        entries = c.fetchall()

        c.execute("SELECT image_path, page_number FROM spec_images WHERE spec_id = ? ORDER BY page_number", (spec['id'],))
        images = c.fetchall()

        results.append({
            'id': spec['id'],
            'spec_number': spec['spec_number'],
            'product_model': spec['product_model'],
            'description': spec['description'],
            'entries': [dict(e) for e in entries],
            'images': [dict(img) for img in images]
        })

    conn.close()

    if results:
        return jsonify({'success': True, 'data': results})
    else:
        return jsonify({'success': False, 'message': '未找到匹配的包装仕样书'})

@app.route('/api/specs', methods=['GET'])
def get_all_specs():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM packaging_specs")
    specs = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': specs})

@app.route('/api/batch_search', methods=['POST'])
def batch_search():
    data = request.get_json()
    items = data.get('items', [])

    if not items:
        return jsonify({'success': False, 'message': '请输入查询内容'})

    conn = get_db()
    c = conn.cursor()

    found = []
    not_found = []

    for item in items:
        customer_name = item.get('customer_name', '').strip()
        valve_model = item.get('valve_model', '').strip()

        if not customer_name or not valve_model:
            not_found.append({
                'customer_name': customer_name,
                'valve_model': valve_model,
                'reason': '输入不完整'
            })
            continue

        c.execute("""
            SELECT DISTINCT ps.id, ps.spec_number, ps.product_model, ps.description
            FROM packaging_specs ps
            JOIN spec_entries se ON ps.id = se.spec_id
            WHERE se.customer_name LIKE ? AND se.valve_model LIKE ?
        """, (f"%{customer_name}%", f"%{valve_model}%"))

        spec = c.fetchone()

        if spec:
            # 获取该仕样书的图片
            c.execute("SELECT image_path FROM spec_images WHERE spec_id = ?", (spec['id'],))
            images = [dict(img)['image_path'] for img in c.fetchall()]
            
            found.append({
                'customer_name': customer_name,
                'valve_model': valve_model,
                'spec_number': spec['spec_number'],
                'product_model': spec['product_model'],
                'description': spec['description'],
                'images': images
            })
        else:
            not_found.append({
                'customer_name': customer_name,
                'valve_model': valve_model,
                'reason': '无该包装仕样书'
            })

    conn.close()

    return jsonify({
        'success': True,
        'found': found,
        'not_found': not_found,
        'total': len(items),
        'found_count': len(found),
        'not_found_count': len(not_found)
    })

@app.route('/api/download_template', methods=['GET'])
def download_template():
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '批量查询模板'

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = '客户名称'
    ws['B1'] = '阀体型号'
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws['B1'].alignment = header_alignment
    ws['A1'].border = thin_border
    ws['B1'].border = thin_border

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    example_data = [
        ('美国三花小包装', 'HDF10HM01'),
        ('东莞三江', 'KDF10H01'),
        ('江苏联孚', 'KDF15H01'),
        ('北京三花制冷', 'KDF10H01'),
    ]

    for i, (cust, model) in enumerate(example_data, start=2):
        ws.cell(row=i, column=1, value=cust).border = thin_border
        ws.cell(row=i, column=2, value=model).border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='批量查询模板.xlsx'
    )

@app.route('/api/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                items.append({
                    'customer_name': str(row[0]).strip(),
                    'valve_model': str(row[1]).strip()
                })

        if not items:
            return jsonify({'success': False, 'message': 'Excel中没有有效数据'})

        conn = get_db()
        c = conn.cursor()

        found = []
        not_found = []

        for item in items:
            customer_name = item['customer_name']
            valve_model = item['valve_model']

            c.execute("""
                SELECT DISTINCT ps.id, ps.spec_number, ps.product_model, ps.description
                FROM packaging_specs ps
                JOIN spec_entries se ON ps.id = se.spec_id
                WHERE se.customer_name LIKE ? AND se.valve_model LIKE ?
            """, (f"%{customer_name}%", f"%{valve_model}%"))

            spec = c.fetchone()

            if spec:
                c.execute("SELECT image_path FROM spec_images WHERE spec_id = ?", (spec['id'],))
                images = [dict(img)['image_path'] for img in c.fetchall()]

                found.append({
                    'customer_name': customer_name,
                    'valve_model': valve_model,
                    'spec_number': spec['spec_number'],
                    'product_model': spec['product_model'],
                    'description': spec['description'],
                    'images': images
                })
            else:
                not_found.append({
                    'customer_name': customer_name,
                    'valve_model': valve_model,
                    'reason': '无该包装仕样书'
                })

        conn.close()

        return jsonify({
            'success': True,
            'found': found,
            'not_found': not_found,
            'total': len(items),
            'found_count': len(found),
            'not_found_count': len(not_found)
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'解析Excel失败: {str(e)}'})

# ============ OCR识别功能 ============

def get_baidu_token():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    
    params = urlencode({
        'grant_type': 'client_credentials',
        'client_id': BAIDU_API_KEY,
        'client_secret': BAIDU_SECRET_KEY
    })
    
    req = urllib.request.Request(BAIDU_TOKEN_URL, params.encode('utf-8'))
    resp = urllib.request.urlopen(req, timeout=10, context=ctx)
    result = json.loads(resp.read())
    return result.get('access_token')

def convert_png_to_jpeg(png_data):
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(png_data))
        if img.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        buf = _io.BytesIO()
        img.save(buf, format='JPEG', quality=90)
        return buf.getvalue()
    except ImportError:
        return png_data

def parse_ocr_result(text_lines):
    result = {
        'spec_number': '',
        'product_model': '',
        'description': '',
        'entries': []
    }
    full_text = '\n'.join(text_lines)
    spec_match = re.search(r'S[-_]?\d+[-_]?BS[-_]?\d+', full_text, re.IGNORECASE)
    if spec_match:
        result['spec_number'] = spec_match.group(0).replace('_', '-')
    model_patterns = [
        r'(HDF[\(（]?\d+[-/]\d+[\)）]?HM)',
        r'(HDF/KDF[\s\S]{0,10}H)',
        r'(KDF\d+[-_]?\d*H\d+)',
        r'(HDF\d+H\d+[A-Z]?)',
    ]
    for pattern in model_patterns:
        model_match = re.search(pattern, full_text, re.IGNORECASE)
        if model_match:
            result['product_model'] = model_match.group(1).strip()
            break
    valve_pattern = re.compile(r'^((?:KDF|HDF)\d+H\d+[A-Z]?)$')
    seq_pattern = re.compile(r'^(\d+)$')
    customer_keywords = [
        '美国三花小包装', '北京三花制冷', '东莞三江', '江苏联孚', '江苏锦东',
        '佛山三枫', '郑州三力', '上海京汉', '成都天奇', '昆山盛年',
        '欧洲三花（小包装）', '欧洲三花', '三花商贸', '深圳市鑫逸生', '盛年制冷设备（昆山）',
        '盛年制冷设备', '成都精工诚志制冷', '沈阳先成', '上海堃霖',
        '上海京汉', '美的集团', '小米', '长虹空调', '天加', '纽恩泰', '芬尼克', '海卓',
        '鑫逸生', '鑫天科技', '申菱', '朗进', '爱美泰', '凯雪', '京鹏',
        '青岛德佰宜', '郑州凯雪运输', '重庆长江', '济南百福特',
        '印度EXCELSIOR', '俄罗斯MORENA', '俄罗斯IMPEX', 'IMPEX'
    ]
    skip_words = [
        '序号', '客户名称', '阀体型号', '特殊包装说明', '备注',
        '浙三司', '技术专用章', 'S-33BS', '附页', '按欧洲', '标贴',
        '阀体型号对照', '产品名称', '包装仕样书', '编制', '校对', '审核',
        '产品型号', '顾客名称', '顾客号', '纸盒', '外箱', '货代码',
        'SANHUA', 'SOLENOID', 'UPC', 'Made', '生产国', '中国',
        '小包装客户清单', '大包装客户清单'
    ]
    entries = []
    current_seq = None
    current_customer = None
    current_valves = []
    for line in text_lines:
        line = line.strip()
        if not line:
            continue
        if any(sw in line for sw in skip_words):
            continue
        if re.match(r'^\d{4}[./]\d{1,2}[./]\d{1,2}', line):
            continue
        if re.match(r'^P/N', line, re.IGNORECASE):
            continue
        if re.match(r'^\d{12,}', line):
            continue
        if seq_pattern.match(line):
            num = int(line)
            if num > 50:
                continue
            if current_seq is not None and current_customer:
                for v in (current_valves if current_valves else ['']):
                    entries.append({
                        'seq_number': current_seq,
                        'customer_name': current_customer,
                        'valve_model': v,
                        'special_note': '',
                        'remark': ''
                    })
            current_seq = num
            current_customer = None
            current_valves = []
            continue
        if valve_pattern.match(line):
            current_valves.append(line)
            continue
        if current_seq is not None and current_customer is None:
            matched = False
            for kw in customer_keywords:
                if kw in line:
                    current_customer = kw
                    matched = True
                    break
            if not matched:
                if not line.startswith('KDF') and not line.startswith('HDF'):
                    current_customer = line
    if current_seq is not None and current_customer:
        for v in (current_valves if current_valves else ['']):
            entries.append({
                'seq_number': current_seq,
                'customer_name': current_customer,
                'valve_model': v,
                'special_note': '',
                'remark': ''
            })
    result['entries'] = entries
    return result

@app.route('/api/ocr', methods=['POST'])
def ocr_recognize():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择图片文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择图片文件'})
    
    try:
        img_data = file.read()
        
        if file.filename.lower().endswith('.png'):
            img_data = convert_png_to_jpeg(img_data)
        
        img_base64 = base64.b64encode(img_data).decode('utf-8')
        
        if len(img_base64) > 4 * 1024 * 1024:
            return jsonify({'success': False, 'message': '图片过大，请压缩后重试'})
        
        token = get_baidu_token()
        if not token:
            return jsonify({'success': False, 'message': '获取百度OCR Token失败'})
        
        ocr_url = BAIDU_OCR_URL + '?access_token=' + token
        post_data = urlencode({'image': img_base64})
        
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        
        req = urllib.request.Request(ocr_url, post_data.encode('utf-8'))
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        
        resp = urllib.request.urlopen(req, timeout=30, context=ctx)
        result = json.loads(resp.read())
        
        if 'words_result' in result:
            text_lines = [item['words'] for item in result['words_result']]
            parsed = parse_ocr_result(text_lines)
            
            return jsonify({
                'success': True,
                'raw_text': text_lines,
                'parsed': parsed
            })
        else:
            error_msg = result.get('error_msg', '未知错误')
            return jsonify({'success': False, 'message': f'OCR识别失败: {error_msg}'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'识别出错: {str(e)}'})

@app.route('/api/ocr_config', methods=['GET'])
def ocr_config():
    return jsonify({
        'success': True,
        'configured': bool(BAIDU_API_KEY and BAIDU_SECRET_KEY)
    })

# ============ 管理后台 ============

@app.route('/admin')
def admin():
    return render_template('admin.html')

@app.route('/api/admin/specs', methods=['GET'])
def admin_get_specs():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT ps.*, 
               GROUP_CONCAT(DISTINCT se.customer_name || ':' || se.valve_model) as entries_summary
        FROM packaging_specs ps
        LEFT JOIN spec_entries se ON ps.id = se.spec_id
        GROUP BY ps.id
        ORDER BY ps.id
    """)
    specs = [dict(row) for row in c.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': specs})

@app.route('/api/admin/spec/<int:spec_id>', methods=['GET'])
def admin_get_spec(spec_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT * FROM packaging_specs WHERE id = ?", (spec_id,))
    spec = c.fetchone()
    if not spec:
        conn.close()
        return jsonify({'success': False, 'message': '仕样书不存在'})

    c.execute("SELECT * FROM spec_entries WHERE spec_id = ? ORDER BY seq_number", (spec_id,))
    entries = [dict(row) for row in c.fetchall()]

    c.execute("SELECT * FROM spec_images WHERE spec_id = ? ORDER BY page_number", (spec_id,))
    images = [dict(row) for row in c.fetchall()]

    conn.close()
    return jsonify({
        'success': True,
        'data': {
            'spec': dict(spec),
            'entries': entries,
            'images': images
        }
    })

@app.route('/api/admin/add_spec', methods=['POST'])
def admin_add_spec():
    data = request.get_json()
    spec_number = data.get('spec_number', '').strip()
    product_model = data.get('product_model', '').strip()
    description = data.get('description', '').strip()
    entries = data.get('entries', [])

    if not spec_number or not product_model:
        return jsonify({'success': False, 'message': '请填写仕样书编号和产品型号'})

    conn = get_db()
    c = conn.cursor()

    c.execute("INSERT INTO packaging_specs (spec_number, product_model, description) VALUES (?, ?, ?)",
              (spec_number, product_model, description))
    spec_id = c.lastrowid

    for entry in entries:
        c.execute("""INSERT INTO spec_entries (spec_id, seq_number, customer_name, valve_model, special_note, remark)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (spec_id, entry.get('seq_number', 0), entry.get('customer_name', ''),
                   entry.get('valve_model', ''), entry.get('special_note', ''), entry.get('remark', '')))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '添加成功', 'spec_id': spec_id})

@app.route('/api/admin/update_spec/<int:spec_id>', methods=['POST'])
def admin_update_spec(spec_id):
    data = request.get_json()
    spec_number = data.get('spec_number', '').strip()
    product_model = data.get('product_model', '').strip()
    description = data.get('description', '').strip()
    entries = data.get('entries', [])

    conn = get_db()
    c = conn.cursor()

    c.execute("UPDATE packaging_specs SET spec_number=?, product_model=?, description=? WHERE id=?",
              (spec_number, product_model, description, spec_id))

    c.execute("DELETE FROM spec_entries WHERE spec_id = ?", (spec_id,))

    for entry in entries:
        c.execute("""INSERT INTO spec_entries (spec_id, seq_number, customer_name, valve_model, special_note, remark)
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (spec_id, entry.get('seq_number', 0), entry.get('customer_name', ''),
                   entry.get('valve_model', ''), entry.get('special_note', ''), entry.get('remark', '')))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '更新成功'})

@app.route('/api/admin/delete_spec/<int:spec_id>', methods=['POST'])
def admin_delete_spec(spec_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("DELETE FROM spec_images WHERE spec_id = ?", (spec_id,))
    c.execute("DELETE FROM spec_entries WHERE spec_id = ?", (spec_id,))
    c.execute("DELETE FROM packaging_specs WHERE id = ?", (spec_id,))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '删除成功'})

@app.route('/api/admin/upload_image', methods=['POST'])
def admin_upload_image():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '请选择文件'})

    file = request.files['file']
    spec_id = request.form.get('spec_id')
    page_number = request.form.get('page_number', 1)

    if file.filename == '':
        return jsonify({'success': False, 'message': '请选择文件'})

    allowed_ext = ('.png', '.jpg', '.jpeg', '.gif', '.bmp')
    if not file.filename.lower().endswith(allowed_ext):
        return jsonify({'success': False, 'message': '请上传图片文件'})

    import time
    import random
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f"spec_{int(time.time())}_{random.randint(1000,9999)}_{page_number}.{ext}"
    filepath = 'static/images/' + filename
    full_path = os.path.join(os.path.dirname(__file__), filepath)

    file.save(full_path)

    if spec_id:
        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO spec_images (spec_id, image_path, page_number) VALUES (?, ?, ?)",
                  (spec_id, filepath, int(page_number)))
        conn.commit()
        conn.close()

    return jsonify({'success': True, 'message': '上传成功', 'image_path': filepath})

@app.route('/api/admin/bind_image', methods=['POST'])
def admin_bind_image():
    data = request.get_json()
    spec_id = data.get('spec_id')
    image_path = data.get('image_path')

    if not spec_id or not image_path:
        return jsonify({'success': False, 'message': '缺少参数'})

    conn = get_db()
    c = conn.cursor()
    c.execute("INSERT INTO spec_images (spec_id, image_path, page_number) VALUES (?, ?, 1)",
              (spec_id, image_path))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '绑定成功'})

@app.route('/api/admin/delete_image/<int:image_id>', methods=['POST'])
def admin_delete_image(image_id):
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT image_path FROM spec_images WHERE id = ?", (image_id,))
    row = c.fetchone()
    if row:
        full_path = os.path.join(os.path.dirname(__file__), row['image_path'])
        if os.path.exists(full_path):
            os.remove(full_path)

    c.execute("DELETE FROM spec_images WHERE id = ?", (image_id,))
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '删除成功'})

if __name__ == '__main__':
    init_db()
    import_data()
    print("=" * 50)
    print("三花包装仕样书查询系统已启动")
    print("请在浏览器中访问: http://localhost:5000")
    print("管理后台: http://localhost:5000/admin")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5000, debug=False)
else:
    # Gunicorn imports this module, so init here too
    init_db()
    import_data()
